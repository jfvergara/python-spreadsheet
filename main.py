import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}  # dictionary
total_value_per_supplier = {}

for products_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(products_row, 4).value
    inventory = product_list.cell(products_row, 2).value
    price = product_list.cell(products_row, 3).value

    # Calculation number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products +1
    else:
        products_per_supplier[supplier_name] = 1

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price





print("Products per supplier:")
print(products_per_supplier)

print("Total value per supplier")
print(total_value_per_supplier)